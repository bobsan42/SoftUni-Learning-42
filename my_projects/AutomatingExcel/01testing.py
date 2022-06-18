import openpyxl as xl
from openpyxl.utils import get_column_letter, column_index_from_string

a = "c:\\Users\\bobsa\\Desktop\\VisualBasic\\export vba.xlsx"
print(a)
wb = xl.load_workbook(a)
# b = xl.open(a)
# print(type(b))
# c.close()
# print (os.getcwd())
# os.chdir('example.xlsx')

print(wb.sheetnames)
wsh = wb['Sheet1']
print('Worksheet %s, maxRow is %s, maxCol is %s' % (wsh, wsh.max_row, wsh.max_column))
print(wb.path)
c = wsh['D3']
# print(c.value)
# Get the row, column, and value from the cell.
print('Row %s, Column %s is %s' % (c.row, c.column, c.value))
'Row 1, Column B is Apples'
print('Cell %s is %s' % (c.coordinate, c.value))
'Cell B1 is Apples'

print(get_column_letter(16385))
print(column_index_from_string("xfd"))
